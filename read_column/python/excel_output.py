#! python

import shutil


class ExcelOputput:
    TemplatePath = r"out/template.xlsx"
    RowHeader = 1
    RowValues = 3
    ColInitial = 2

    def __init__(self, path):
        self.path = path

    def write(self, nkc):
        from openpyxl import load_workbook

        shutil.copyfile(self.TemplatePath, self.path)
        wb = load_workbook(filename=self.path)
        ws = wb['Sheet1']

        self.write_header(ws, nkc)
        self.write_values(ws, nkc)

        wb.save(self.path)
        return self

    def write_header(self, ws, nkc):
        header_data = self.setup_header_data(nkc)
        for irow, row in enumerate(header_data):
            for icol, val in enumerate(row):
                nrow = irow+self.RowHeader
                ncol = icol+self.ColInitial
                ws.cell(row=nrow, column=ncol).value = val
        return self

    def setup_header_data(self, nkc):
        row_name_key = []
        row_category = []
        for cat in nkc.categories:
            name_key = "{} & {}".format(nkc.header.name, nkc.header.key)
            row_name_key.append(name_key)
            row_name_key.append("")
            row_category.append(cat.name)
            row_category.append("")

        header_data = []
        header_data.append(row_name_key)
        header_data.append(row_category)

        return header_data

    def write_values(self, ws, nkc):
        values_data = self.setup_values_data(nkc)
        for irow, row in enumerate(values_data):
            for icol, val in enumerate(row):
                ws.cell(row=irow+self.RowValues, column=icol+self.ColInitial)
        return self

    def setup_values_data(self, nkc):
        row_max = max([len(cat.xys) for cat in nkc.categories])
        values_data = []
        for irow in range(row_max):
            row_data = []
            for cat in nkc.categories:
                if irow < len(cat.xys):
                    row_data.append(cat.xys[irow].x)
                    row_data.append(cat.xys[irow].y)
                else:
                    row_data.append("")
                    row_data.append("")
            values_data.append(row_data)
        return values_data


if __name__ == '__main__':
    from excel_input import ExcelInput

    path = r'inp/data1.xlsx'
    nkc_list = ExcelInput(path).to_NKCs()

    for nkc in nkc_list:
        print(nkc.excel_output_path())
        nkc.write_excel_output()
